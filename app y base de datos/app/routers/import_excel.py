"""
Excel Import Router for Products
Handles importing products from Excel files with Spanish columns
"""
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from sqlalchemy.orm import Session
from typing import List, Optional
import io
import re

from ..database import get_db
from ..models import Product, Category
from ..auth_utils import get_current_user, User
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

router = APIRouter(prefix="/import", tags=["Import"])


def parse_colombian_price(price_str: str) -> float:
    """
    Parse Colombian price format (e.g., "$ 28.000" or "28.000") to float
    Colombian format uses . as thousands separator
    """
    if not price_str:
        return 0.0
    
    # Remove currency symbol, spaces, and convert
    cleaned = str(price_str).replace('$', '').replace(' ', '').strip()
    
    # If it contains a comma, it might be decimal separator
    if ',' in cleaned and '.' in cleaned:
        # Format: 1.234,56 (European/Colombian with decimals)
        cleaned = cleaned.replace('.', '').replace(',', '.')
    elif '.' in cleaned:
        # Check if it's thousands separator (no decimals) or decimal
        # If more than 2 digits after last dot, it's thousands separator
        parts = cleaned.split('.')
        if len(parts[-1]) == 3:
            # It's thousands separator: 28.000 -> 28000
            cleaned = cleaned.replace('.', '')
        # else it's already decimal format
    
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def normalize_header(header: str) -> str:
    """Normalize header name for matching"""
    if not header:
        return ""
    return (header.lower()
            .strip()
            .replace('á', 'a').replace('é', 'e')
            .replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
            .replace('.', '')
            .replace(' ', '_'))


@router.post("/excel")
async def import_excel(
    file: UploadFile = File(...),
    category_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Import products from Excel file.
    
    Expected columns (Spanish):
    - NOMBRE: Product name (required)
    - EXISTENCIA: Stock quantity (required)
    - P. COSTO / P_COSTO / PRECIO_COSTO: Cost price
    - P. VENTA / P_VENTA / PRECIO_VENTA: Sale price  
    - PROVEEDOR: Supplier name
    - MARCA: Brand
    - MODELO: Model
    """
    
    # Check file extension
    ext = file.filename.lower()
    if not ext.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos Excel (.xlsx, .xls) o CSV (.csv)")
    
    rows_data = []
    headers = []
    
    try:
        content = await file.read()
        
        # Handle CSV
        if ext.endswith('.csv'):
            import csv
            # Try to decode with utf-8-sig (common for Excel CSVs)
            try:
                decoded = content.decode('utf-8-sig')
            except UnicodeDecodeError:
                decoded = content.decode('latin-1')
            
            # Detect delimiter
            delimiter = ';' if ';' in decoded.splitlines()[0] else ','
            
            f = io.StringIO(decoded)
            reader = csv.reader(f, delimiter=delimiter)
            
            # Read all rows
            all_rows = list(reader)
            if not all_rows:
                raise HTTPException(status_code=400, detail="El archivo CSV está vacío")
                
            # Headers are first row
            raw_headers = all_rows[0]
            for h in raw_headers:
                headers.append(normalize_header(str(h)))
            
            # Data rows (store as dict-like or list to match existing logic??)
            # Existing logic iterates sheet rows. Let's make rows_data a list of lists 
            # similar to what openpyxl returns (values only)
            rows_data = all_rows[1:]
            
        # Handle Excel
        else:
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise HTTPException(status_code=500, detail="Módulo openpyxl no instalado")
            
            workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            
            # Get headers from first row
            for cell in sheet[1]:
                headers.append(normalize_header(str(cell.value) if cell.value else ""))
            
            # Get all rows
            rows_data = list(sheet.iter_rows(min_row=2, values_only=True))
            
        # Map column names to indices
        column_map = {
            'nombre': None,
            'existencia': None,
            'stock': None,
            'p_costo': None,
            'precio_costo': None,
            'costo': None,
            'p_venta': None, 
            'precio_venta': None,
            'venta': None,
            'precio': None,
            'proveedor': None,
            'marca': None,
            'modelo': None,
            'descripcion': None,
        }
        
        for idx, header in enumerate(headers):
            for key in column_map.keys():
                if key in header or header in key:
                    column_map[key] = idx
                    break
        
        # Determine which columns to use
        name_col = column_map.get('nombre')
        stock_col = column_map.get('existencia') or column_map.get('stock')
        cost_col = column_map.get('p_costo') or column_map.get('precio_costo') or column_map.get('costo')
        price_col = column_map.get('p_venta') or column_map.get('precio_venta') or column_map.get('venta') or column_map.get('precio')
        supplier_col = column_map.get('proveedor')
        brand_col = column_map.get('marca')
        model_col = column_map.get('modelo')
        desc_col = column_map.get('descripcion')
        
        # New: Category column detection
        cat_col_idx = None
        for idx, header in enumerate(headers):
            if 'categoria' in header or 'category' in header:
                cat_col_idx = idx
                break
        
        if name_col is None:
            raise HTTPException(
                status_code=400, 
                detail=f"No se encontró columna 'NOMBRE'. Columnas encontradas: {headers}"
            )
        


        # Pre-fetch categories for faster lookup
        result = await db.execute(select(Category))
        categories_db = result.scalars().all()
        categories_map = {c.name.lower(): c.id for c in categories_db}
        
        # Process rows
        products_created = 0
        products_updated = 0
        errors = []
        
        for row_idx, row in enumerate(rows_data, start=2):
            try:
                # Skip empty rows
                if not row or not row[name_col]:
                    continue
                
                name = str(row[name_col]).strip()
                if not name:
                    continue
                
                # Determine category for this row
                row_category_id = category_id
                
                # If no global category selected, try to find in row
                if row_category_id is None and cat_col_idx is not None and row[cat_col_idx]:
                    cat_name = str(row[cat_col_idx]).lower().strip()
                    
                    # Synonym mapping
                    synonyms = {
                        'iphone': 'apple',
                        'ipad': 'apple',
                        'mac': 'apple',
                        'watch': 'accesorios',
                        'funda': 'accesorios',
                        'vidrio': 'accesorios',
                        'cargador': 'accesorios',
                        'samsung': 'android',
                        'xiaomi': 'android',
                        'redmi': 'android',
                        'motorola': 'android',
                        'play': 'cacharros',
                        'consola': 'cacharros'
                    }
                    
                    # Check synonyms
                    if cat_name in synonyms:
                        cat_name = synonyms[cat_name]
                    
                    # Try direct match
                    if cat_name in categories_map:
                        row_category_id = categories_map[cat_name]
                    # Try partial match (e.g. "accesorios para celulares" -> "accesorios")
                    else:
                        for db_cat_name, db_cat_id in categories_map.items():
                            if db_cat_name in cat_name or cat_name in db_cat_name:
                                row_category_id = db_cat_id
                                break
                
                # Parse values
                stock = int(row[stock_col]) if stock_col is not None and row[stock_col] else 0
                cost_price = parse_colombian_price(str(row[cost_col])) if cost_col is not None and row[cost_col] else 0
                sale_price = parse_colombian_price(str(row[price_col])) if price_col is not None and row[price_col] else 0
                
                # Fallback: if sale price is 0/None, use cost price
                if not sale_price and cost_price:
                    sale_price = cost_price
                
                supplier = str(row[supplier_col]).strip() if supplier_col is not None and row[supplier_col] else None
                brand = str(row[brand_col]).strip() if brand_col is not None and row[brand_col] else None
                model = str(row[model_col]).strip() if model_col is not None and row[model_col] else None
                description = str(row[desc_col]).strip() if desc_col is not None and row[desc_col] else None
                
                # Check if product exists (by name)
                result = await db.execute(select(Product).where(Product.name == name))
                existing = result.scalar_one_or_none()
                
                if existing:
                    # Update existing product
                    existing.stock = stock
                    if cost_price:
                        existing.cost_price = cost_price
                    if sale_price:
                        existing.price = sale_price
                    if supplier:
                        existing.supplier = supplier
                    if brand:
                        existing.brand = brand
                    if model:
                        existing.model = model
                    if row_category_id:
                        existing.category_id = row_category_id
                    products_updated += 1
                else:
                    # Create new product
                    new_product = Product(
                        name=name,
                        stock=stock,
                        cost_price=cost_price,
                        price=sale_price,
                        supplier=supplier,
                        brand=brand,
                        model=model,
                        description=description,
                        category_id=row_category_id,
                        condition="new",
                        is_active=True
                    )
                    db.add(new_product)
                    products_created += 1
                    
            except Exception as e:
                errors.append(f"Fila {row_idx}: {str(e)}")
                continue
        
        await db.commit()
        
        return {
            "success": True,
            "message": f"Importación completada",
            "products_created": products_created,
            "products_updated": products_updated,
            "errors": errors[:10] if errors else [],
            "total_errors": len(errors)
        }
        
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Error procesando archivo: {str(e)}")
